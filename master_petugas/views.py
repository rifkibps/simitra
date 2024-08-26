import json
from django.urls import reverse, reverse_lazy
from django.core.serializers.json import DjangoJSONEncoder

from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.shortcuts import render
from django.views import View
from django.core import serializers
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from . import models, forms, utils

import itertools

from .resources import MasterPetugasResource, MasterAlokasiResource, MasterRoleResource
from master_survey.models import SurveyModel
from master_penilaian.models import MasterNilaiPetugas, KegiatanPenilaianModel

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill
from . import utils


from statistics import mean
from operator import itemgetter
from django.shortcuts import redirect

from pprint import pprint 
# Packages for upload master petugas

# Create your views here.

class MasterPetugasJsonResponseClassView(LoginRequiredMixin, View):

    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)


        data = models.MasterPetugas.objects
        if datatables.get('pendidikan_filter'):
            data = data.filter(pendidikan = datatables.get('pendidikan_filter'))

        if datatables.get('status_filter'):
            data = data.filter(status = datatables.get('status_filter'))

        data = data.all().exclude(Q(kode_petugas=None)|Q(nama_petugas=None)|Q(nik=None)|Q(email=None)|Q(no_telp=None)|Q(status=None))
        records_total = data.count()
        records_filtered = records_total
        
        if search:

            data = models.MasterPetugas.objects

            if datatables.get('pendidikan_filter'):
                data = data.filter(pendidikan = datatables.get('pendidikan_filter'))

            if datatables.get('status_filter'):
                data = data.filter(status = datatables.get('status_filter'))

            data = data.filter(
                Q(kode_petugas__icontains=search)|
                Q(nama_petugas__icontains=search)|
                Q(nik__icontains=search)|
                Q(email__icontains=search)|
                Q(no_telp__icontains=search)|
                Q(status__icontains=search)
            ).exclude(Q(nama_petugas=None)|Q(nik=None)|Q(email=None)|Q(no_telp=None)|Q(status=None))

          
            records_total = data.count()
            records_filtered = records_total
        
        data = data.order_by(order_col_name)
        # Conf Paginator
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        data = []

        for obj in object_list:
            
            if obj.status == '0':
                class_badge = 'badge badge-success-lighten'
            elif obj.status == '1':
                class_badge = 'badge badge-secondary-lighten'
            elif obj.status == '3':
                class_badge = 'badge badge-danger-lighten'
            else:
                class_badge = 'badge badge-primary-lighten'

            data.append({
                        'kode_petugas': obj.kode_petugas,
                        'nama_petugas': f'<a href="{reverse_lazy("master_petugas:mitra-view-detail", kwargs={"mitra_id": obj.id})}" class="text-body">{obj.nama_petugas}</a>',
                        'nik': obj.nik,
                        'email': obj.email,
                        'no_telp': obj.no_telp,
                        'status': f'<span class="badge {class_badge}"> {obj.get_status_display()} </span>',
                        'aksi': f'<a href="javascript:void(0);" onclick="editPetugas({obj.id})" class="action-icon"><i class="mdi mdi-square-edit-outline"></i></a> <a href="javascript:void(0);" onclick="hapusPetugas({obj.id});" class="action-icon"> <i class="mdi mdi-delete"></i></a>'
                })
            
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }


class MasterPetugasClassView(LoginRequiredMixin, View):

    data_petugas = models.MasterPetugas.objects

    def post(self, request):

        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.MasterPetugasForm(request.POST)

            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)


    def get(self, request):

        form = forms.MasterPetugasForm()
        context = {
            'title' : 'Master Mitra',
            'data_petugas' : self.data_petugas.all(),
            'form': form,
            'form_upload': forms.MasterPetugasFormUpload()
            }

        return render(request, 'master_petugas/index.html', context)


class MasterPetugasDeleteView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')

                data_petugas = models.MasterPetugas.objects.filter(pk = id)
                if data_petugas.exists():

                    check_nilai_mitra = MasterNilaiPetugas.objects.filter(petugas__petugas = id)
                    if check_nilai_mitra.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data alokasi petugas telah digunakan pada master data penilaian'}, status=200)

                    data_petugas.delete()
                    return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)


class MasterPetugasDetailView(LoginRequiredMixin, View):


    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')
                data_petugas = models.MasterPetugas.objects.filter(pk=id)

                if data_petugas.exists():
                    return JsonResponse({'status' : 'success', 'instance': list(data_petugas.values())[0]}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 


class MasterPetugasUpdateView(LoginRequiredMixin, View):
    data_petugas = models.MasterPetugas.objects

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:

            data = get_object_or_404(models.MasterPetugas, pk=request.POST.get('id'))

            form = forms.MasterPetugasForm(request.POST, instance=data)

            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)


class MasterPetugasExportClassView(LoginRequiredMixin, View):
    def get(self, request):
    
        resource = MasterPetugasResource()
        dataset = resource.export()

        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Master Petugas.xls"'
        return response 
    

class MasterPetugasTemplateClassView(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs): 

        def_rows = self.kwargs['rows']
        wb = Workbook()

        ws = wb.active

        # Ini untuk header columns
        ws.title = 'Upload Data Petugas'

        header = utils.get_verbose_fields(models.MasterPetugas, exclude_pk=True)
        header = ['No'] + header

        head_row = 2
        header_cols = np.array(ws[f'A{head_row}':f'M{head_row}'])

        # Set value and style for header
        for v,c in zip(header, header_cols.T.flatten()):
            # Set style
            c.font = Font(name='Cambria', size=12)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type = "solid")
            c.value = v

        # Adjustment cols
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
            
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 22.50


        utils.generate_meta_templates(ws, 'P', 3, 'Data Pendidikan', list(models.MasterPetugas._meta.get_field('pendidikan').choices), 'G', 3, def_rows=def_rows)
        utils.generate_meta_templates(ws, 'R', 3, 'Agama', list(models.MasterPetugas._meta.get_field('agama').choices), 'I',3,  def_rows=def_rows)
        utils.generate_meta_templates(ws, 'T', 3, 'Status Mitra', list(models.MasterPetugas._meta.get_field('status').choices), 'M', 3,  def_rows=def_rows)

        ws.merge_cells('A1:M1')
        ws['A1'] = 'Template Upload Data Mitra'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Cambria',bold=True, size=14)

        ws.merge_cells('P2:T2')
        ws['P2'] = 'Metadata Mitra'
        ws['P2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['P2'].font = Font(name='Cambria',bold=True, size=12)

        for col in ['P', 'Q', 'R', 'S', 'T']:
            ws.column_dimensions[col].hidden= True

        for row in range(def_rows):
            ws[f'A{row+3}'] = row+1
            ws[f'A{row+3}'].alignment = Alignment(horizontal='center', vertical='center')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Template Upload Data Mitra.xlsx'

        wb.save(response)
        return response


class MasterPetugasUploadClassView(LoginRequiredMixin, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
           
            form = forms.MasterPetugasFormUpload(request.POST, request.FILES)
            model = models.MasterPetugas

            if form.is_valid():
                
                df = form.cleaned_data
                objs = []
                for idx in range(len(df['id'])):

                    objs.append(
                        model(
                            kode_petugas = df['kode_petugas'][idx],
                            nama_petugas= df['nama_petugas'][idx],
                            nik= df['nik'][idx],
                            npwp= df['npwp'][idx],
                            tgl_lahir= df['tgl_lahir'][idx],
                            pendidikan= df['pendidikan'][idx],
                            pekerjaan= df['pekerjaan'][idx],
                            agama= df['agama'][idx],
                            email= df['email'][idx],
                            no_telp= df['no_telp'][idx],
                            alamat= df['alamat'][idx],
                            status = df['status'][idx],
                        )
                    )

                model.objects.bulk_create(objs)
                return JsonResponse({"status": "success", "messages": f"<strong>{len(df['id'])}</strong> Data berhasil diupload."})
            else:
                
                error_messages = list(itertools.chain.from_iterable(form.errors['import_file'].as_data()))
                return JsonResponse({"status": "error", "messages": error_messages})

        return JsonResponse({"error": ""}, status=403)  


class MasterPetugasDetailViewClassView(LoginRequiredMixin, View):
    
    def _globalRank(self, request):
        data = MasterNilaiPetugas.objects.values('petugas__petugas__kode_petugas', 'petugas__petugas__nama_petugas', 'petugas__survey__nama', 'petugas__role__jabatan', 'penilaian__kegiatan_penilaian','penilaian__kegiatan_penilaian__nama_kegiatan', 'nilai', 'catatan')
        
        master_data = []
        
        for dt in data:
  
            check_exist = [index for (index, d) in enumerate(master_data) if d["kode_petugas"] == dt['petugas__petugas__kode_petugas']]
            
            if len(check_exist) > 0:

                check_exist_2 = [index for (index, d) in enumerate(master_data[check_exist[0]]['kegiatan_penilaian']) if d["id_kegiatan"] == dt['penilaian__kegiatan_penilaian']]
                
                if len(check_exist_2) > 0:
                    master_data[check_exist[0]]['kegiatan_penilaian'][check_exist_2[0]]['nilai'].append(dt['nilai'])
                    master_data[check_exist[0]]['kegiatan_penilaian'][check_exist_2[0]]['catatan'].append(dt['catatan'])
                else:
                    master_data[check_exist[0]]['kegiatan_penilaian'].append({
                        'id_kegiatan' : dt['penilaian__kegiatan_penilaian'],
                        'survey' : dt['petugas__survey__nama'],
                        'nama_kegiatan': dt['penilaian__kegiatan_penilaian__nama_kegiatan'],
                        'role': dt['petugas__role__jabatan'],
                        'nilai' : [dt['nilai']],
                        'catatan' : [dt['catatan']],
                    })

                continue

            master_data.append({
                'kode_petugas': dt['petugas__petugas__kode_petugas'],
                'nama_petugas': dt['petugas__petugas__nama_petugas'],
                'rerata_final': 0,
                'ranking_final': 0,
                'kegiatan_penilaian' : [{'id_kegiatan': dt['penilaian__kegiatan_penilaian'] , 'role' :  dt['petugas__role__jabatan'], 'survey' : dt['petugas__survey__nama'], 'nama_kegiatan': dt['penilaian__kegiatan_penilaian__nama_kegiatan'], 'nilai': [dt['nilai']], 'catatan': [dt['catatan']]}]
            })

        for dt in master_data:
            
            mean_data = []
            for dt_kegiatan in dt['kegiatan_penilaian']:
                mean_data.append(round(mean(dt_kegiatan['nilai']), 2))
        
            dt['rerata_final'] = round(mean(mean_data), 2)

        data_sorted = sorted(master_data, key = itemgetter('rerata_final'), reverse=True)
        for idx, dt in enumerate(data_sorted):
            dt['ranking_final'] = idx+1

        return data_sorted


    def get(self, request, *args, **kwargs):
        
        mitra_id = self.kwargs['mitra_id']
        mitra = models.MasterPetugas.objects.filter(pk = mitra_id)

        if mitra.exists() == False:
            return redirect(request.META.get('HTTP_REFERER', '/'))

        survei_ = models.AlokasiPetugas.objects.filter(petugas = mitra_id)

        kegiatan_penilaian_ = MasterNilaiPetugas.objects.filter(petugas__petugas = mitra_id).values('penilaian__kegiatan_penilaian')

        global_rank = self._globalRank(request)

        # Mengurutkan penilaian berdasarkan kegiatan penilaian
        data_nilai_mitra = {}
        for dt in global_rank:
            dt_kode_petugas = dt['kode_petugas']
            dt_kegiatan_penilaian = dt['kegiatan_penilaian']

            for dt_ in dt_kegiatan_penilaian:

                dt_['rerata'] = round(mean(dt_['nilai']), 2)

                if dt_['id_kegiatan'] in data_nilai_mitra:
                    data_nilai_mitra[dt_['id_kegiatan']].append(dt_ | {'kode_petugas' : dt_kode_petugas})
                else:
                    data_nilai_mitra[dt_['id_kegiatan']] = [
                        dt_ | {'kode_petugas' : dt_kode_petugas}
                    ]

        for idx, val in data_nilai_mitra.items():
            sorted_ = sorted(val, key = itemgetter('rerata'), reverse=True)
            
            for idx2, dt_sort in enumerate(sorted_):
                dt_sort['rank'] = f'{idx2+1} of {len(sorted_)}'

            data_nilai_mitra[idx] = sorted_


        # Formatting Data
        data_final = []
        for dt_ in kegiatan_penilaian_.distinct():
            id_kegiatan_penilaian = dt_['penilaian__kegiatan_penilaian']
            if id_kegiatan_penilaian in data_nilai_mitra:
                filter_data = [index for (index, d) in enumerate(data_nilai_mitra[id_kegiatan_penilaian]) if d["kode_petugas"] == mitra.first().kode_petugas]

                if len(filter_data) > 0:
                    data_final.append(data_nilai_mitra[id_kegiatan_penilaian][filter_data[0]])


        for dt in data_final:
            dt['catatan'] = np.unique(np.array(dt['catatan']))
        
        check_exist = [index for (index, d) in enumerate(global_rank) if d["kode_petugas"] == mitra.first().kode_petugas]
        global_ranking = global_rank[check_exist[0]]['ranking_final'] if len(check_exist) > 0 else ''

        context = {
            'title' : f'{mitra.first().kode_petugas} | {mitra.first().nama_petugas}',
            'mitra' : mitra.first(),
            'survei_followed' : survei_.count(),
            'history_survey' : survei_,
            'kegiatan_followed' : kegiatan_penilaian_.distinct().count(),
            'global_rank' : global_ranking,
            'penilaian' : data_final
        }

        return render(request, 'master_petugas/detail_petugas.html', context)

class MasterPetugasSearchClassView(LoginRequiredMixin, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:

            print('MASUK')

            if request.method == 'POST':
                
                search_by = request.POST.get('search_by')
                data_petugas = models.MasterPetugas.objects.filter(
                    Q(kode_petugas__icontains=search_by)|
                    Q(nama_petugas__icontains=search_by)|
                    Q(nik__icontains=search_by)|
                    Q(email__icontains=search_by)|
                    Q(no_telp__icontains=search_by)
                ).exclude(Q(nama_petugas=None)|Q(nik=None)|Q(email=None)|Q(no_telp=None)|Q(status=None))

            row_counts = data_petugas.count()
            search_result = ''
            for dt in data_petugas[:5]:
                search_result += f'<a href="{reverse_lazy("master_petugas:mitra-view-detail", kwargs={"mitra_id": dt.id})}" class="dropdown-item notify-item"><i class="mdi mdi-account-circle-outline me-1"></i><span>{dt.nama_petugas} ({dt.kode_petugas})</span></a>'
          
            return JsonResponse({'status' : 'success', 'search_result': {'row_count' : row_counts , 'search_result' : search_result }}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 




### Alokasi Petugas

class AlokasiPetugasClassView(LoginRequiredMixin, View):
    
    def get(self, request):
        context = {
            'title' : 'Alokasi Mitra',
            'data_mitra': models.MasterPetugas.objects.filter(~Q(status = 3), ~Q(status = 1)),
            'data_survei' : SurveyModel.objects.all(),
            'data_jabatan' : models.RoleMitra.objects.all(),
            'form_upload' : forms.AlokasiPetugasFormUpload(),
            'form' : forms.AlokasiForm()
            }

        return render(request, 'master_petugas/alokasi.html', context)


    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            form = forms.AlokasiForm(request.POST)

            if form.is_valid():

                data = get_object_or_404(models.MasterPetugas, pk=form.data['petugas'])
                if data.status == '1' or data.status == '3':
                    return JsonResponse({"status":"failed", 'message': f'Mitra dengan status {data.get_status_display()} tidak dapat dialokasikan.'}, status=200)

                check_exist_data = models.AlokasiPetugas.objects.filter(petugas = form.data['petugas'], survey = form.data['survey'], role = form.data['role'])
                
                if check_exist_data.exists():
                    return JsonResponse({"status":"failed", 'message': 'Data alokasi telah tersedia pada database'}, status=200)

                instance = form.save()
                user_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"status":"success", 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                return JsonResponse({"status":"failed", "error": form.errors}, status=400)
        return JsonResponse({"status":"failed", "error": ""}, status=400)
    

class AlokasiPetugasDeleteView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')

                data_petugas = models.AlokasiPetugas.objects.filter(pk = id)
                if data_petugas.exists():

                    check_nilai_mitra = MasterNilaiPetugas.objects.filter(petugas = id)

                    if check_nilai_mitra.exists() == False:
                        data_petugas.delete()
                        return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
                    
                    return JsonResponse({'status' : 'failed', 'message': 'Data alokasi petugas telah digunakan pada master data penilaian.'}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)


class MasterAlokasiDetailView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')
                data_petugas = models.AlokasiPetugas.objects.filter(pk=id)

                check_nilai_mitra = MasterNilaiPetugas.objects.filter(petugas = id)

                if check_nilai_mitra.exists():
                    return JsonResponse({'status' : 'failed', 'message': 'Data alokasi petugas telah digunakan pada master data penilaian'}, status=200)
                
                if data_petugas.exists():
                    return JsonResponse({'status' : 'success', 'instance': list(data_petugas.values())[0]}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 


class MasterAlokasiUpdateView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:

            data = get_object_or_404(models.AlokasiPetugas, pk=request.POST.get('id'))

            if data.petugas.status == '1' or data.petugas.status == '3':
                return JsonResponse({"status":"failed", 'message': f'Mitra dengan status {data.petugas.get_status_display()} tidak dapat dialokasikan.'}, status=200)
        
            check_nilai_mitra = MasterNilaiPetugas.objects.filter(petugas = request.POST.get('id'))
            if check_nilai_mitra.exists():
                return JsonResponse({'status' : 'failed', 'message': 'Data alokasi petugas telah digunakan pada master data penilaian'}, status=200)

            form = forms.AlokasiForm(request.POST, instance=data)

            if form.is_valid():
            
                check_exist_data = models.AlokasiPetugas.objects.filter(~Q(pk=form.data['id']) & Q(petugas = form.data['petugas']) & Q(survey = form.data['survey']) & Q(role = form.data['role']))
              
                if check_exist_data.exists():
                    return JsonResponse({"status":"failed", 'message': 'Data alokasi telah tersedia pada database'}, status=200)

                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"status":"success", "instance": ser_instance, "message": "Data berhasil diubah"}, status=200)
            else:
                return JsonResponse({"status": "failed", "error": form.errors, "message": "Terjadi kesalahan"}, status=400)
            
        return JsonResponse({"status": "failed", "message": "Terjadi Kesalahan"}, status=400)


class MasterAlokasiJsonResponseClassView(LoginRequiredMixin, View):

    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.AlokasiPetugas.objects
        if datatables.get('survei_filter'):
            data = data.filter(survey = datatables.get('survei_filter'))

        if datatables.get('jabatan_filter'):
            data = data.filter(role = datatables.get('jabatan_filter'))

        data = data.all().exclude(Q(petugas=None)|Q(role=None)|Q(survey=None))
        records_total = data.count()
        records_filtered = records_total
        
        if search:
            data = models.AlokasiPetugas.objects.filter(
                Q(petugas__kode_petugas__icontains=search)|
                Q(petugas__nama_petugas__icontains=search)|
                Q(survey__nama__icontains=search)|
                Q(role__jabatan__icontains=search)
            ).exclude(Q(petugas=None)|Q(role=None)|Q(survey=None))

            records_total = data.count()
            records_filtered = records_total
        
        data = data.order_by(order_col_name)
        # Conf Paginator
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        data = [
            {
                'petugas__kode_petugas': obj.petugas.kode_petugas,
                'petugas__nama_petugas': f'<a href="{reverse_lazy("master_petugas:mitra-view-detail", kwargs={"mitra_id": obj.id})}" class="text-body" target="_blank">{obj.petugas.nama_petugas}</a>',
                'survey__nama': obj.survey.nama,
                'role__jabatan': obj.role.jabatan,
                'aksi': f'<a href="javascript:void(0);" onclick="editAlokPetugas({obj.id})" class="action-icon"><i class="mdi mdi-square-edit-outline"></i></a> <a href="javascript:void(0);" onclick="deleteAlokasi({obj.id});" class="action-icon"> <i class="mdi mdi-delete"></i></a>'
                
            } for obj in object_list
        ]
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }


class MasterAlokasiExportClassView(LoginRequiredMixin, View):
    def get(self, request):
    
        resource = MasterAlokasiResource()
        dataset = resource.export()

        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Alokasi Petugas.xls"'
        return response 

class MasterAlokasiTemplateClassView(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs): 

        def_rows = self.kwargs['rows']
        wb = Workbook()

        ws = wb.active

        # Ini untuk header columns
        ws.title = 'Upload Alokasi Petugas'

        header = utils.get_verbose_fields(models.AlokasiPetugas, exclude_pk=True)
        header = ['No'] + header

        head_row = 2
        header_cols = np.array(ws[f'A{head_row}':f'M{head_row}'])

        # Set value and style for header
        for v,c in zip(header, header_cols.T.flatten()):
            # Set style
            c.font = Font(name='Cambria', size=12)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type = "solid")
            c.value = v

        # Adjustment cols
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
            
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 22.50

        utils.generate_meta_templates_multiple_cols(ws, 'G', 3, 'Data Mitra', list(models.MasterPetugas.objects.filter(~Q(status=1), ~Q(status=3)). values_list('kode_petugas', 'nama_petugas')), 'B', 3, def_rows = def_rows)
        utils.generate_meta_templates(ws, 'J', 3, 'Data Survei/Sensus', list(SurveyModel.objects.values_list('id','nama')), 'C',3, def_rows=def_rows)
        utils.generate_meta_templates(ws, 'L', 3, 'Jabatan Mitra', list(models.RoleMitra.objects.values_list('id','jabatan')), 'D', 3, def_rows=def_rows)

        ws.merge_cells('A1:D1')
        ws['A1'] = 'Template Upload Alokasi Mitra'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Cambria',bold=True, size=14)

        ws.merge_cells('G2:L2')
        ws['G2'] = 'Metadata Alokasi Mitra'
        ws['G2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['G2'].font = Font(name='Cambria',bold=True, size=12)

        for row in range(def_rows):
            ws[f'A{row+3}'] = row+1
            ws[f'A{row+3}'].alignment = Alignment(horizontal='center', vertical='center')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Template Upload Alokasi Mitra.xlsx'

        wb.save(response)
        return response


class MasterAlokasiUploadClassView(LoginRequiredMixin, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
           
            form = forms.AlokasiPetugasFormUpload(request.POST, request.FILES)
            model = models.AlokasiPetugas
            model_mitra = models.MasterPetugas
            model_survey = SurveyModel
            model_role = models.RoleMitra

            if form.is_valid():
                
                df = form.cleaned_data
                objs = []
                for idx in range(len(df['id'])):

                    objs.append(
                        model(
                            petugas = model_mitra.objects.get(pk = df['petugas'][idx]),
                            survey=  model_survey.objects.get(pk = df['survey'][idx]),
                            role = model_role.objects.get(pk = df['role'][idx])
                        )
                    )

                model.objects.bulk_create(objs)
                return JsonResponse({"status": "success", "messages": f"<strong></strong> Data berhasil diupload."})
            else:
                
                error_messages = list(itertools.chain.from_iterable(form.errors['import_file'].as_data()))
                return JsonResponse({"status": "error", "messages": error_messages})

        return JsonResponse({"error": ""}, status=403)  


# Role Petugas

class RolePetugasClassView(LoginRequiredMixin, View):
    
    def get(self, request):
        context = {
            'title' : 'Role Petugas',
            'form' : forms.RoleForm()
            }

        return render(request, 'master_petugas/role.html', context)


    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.RoleForm(request.POST)
            if form.is_valid():
                instance = form.save()
                user_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": user_instance, 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)
    

class MasterRoleJsonResponseClassView(LoginRequiredMixin, View):

    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for 
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.RoleMitra.objects.all().exclude(Q(jabatan=None))
        records_total = data.count()
        records_filtered = records_total
        
        if search:
            data = models.RoleMitra.objects.filter(
                Q(jabatan__icontains=search)
            ).exclude(Q(jabatan=None))

            records_total = data.count()
            records_filtered = records_total
        
        data = data.order_by(order_col_name)
        # Conf Paginator
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        data = [
            {
                'jabatan': obj.jabatan,
                'aksi': f'<button onclick="editRole({obj.id})" class="btn btn-primary"><i class="mdi mdi-square-edit-outline"></i></button> <button onclick="deleteRole({obj.id});" class="btn btn-danger"> <i class="mdi mdi-delete"></i></button>',
            } for obj in object_list
        ]
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }


class RolePetugasDeleteView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')

                data_petugas = models.RoleMitra.objects.filter(pk = id)
                if data_petugas.exists():
                    
                    check_kegiatan_penilaian = KegiatanPenilaianModel.objects.filter(role_permitted = id)
                    if check_kegiatan_penilaian.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data role petugas telah digunakan pada master data penilaian'}, status=200)
                
                    check_nilai_mitra = MasterNilaiPetugas.objects.filter(petugas__role = id)
                    if check_nilai_mitra.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data role petugas telah digunakan pada master data penilaian'}, status=200)

                    data_petugas.delete()
                    return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)

class MasterRoleDetailView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')
                role = models.RoleMitra.objects.filter(pk=id)

                if role.exists():
                    
                    check_kegiatan_penilaian = KegiatanPenilaianModel.objects.filter(role_permitted = id)
                    if check_kegiatan_penilaian.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data role petugas telah digunakan pada master data penilaian'}, status=200)
                    
                    check_nilai_mitra = MasterNilaiPetugas.objects.filter(petugas__role = id)
                    if check_nilai_mitra.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data role petugas telah digunakan pada master data penilaian'}, status=200)

                    return JsonResponse({'status' : 'success', 'instance': list(role.values())[0]}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 


class MasterRoleUpdateView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:

            data = get_object_or_404(models.RoleMitra, pk=request.POST.get('id'))

            check_kegiatan_penilaian = KegiatanPenilaianModel.objects.filter(role_permitted = request.POST.get('id'))
            if check_kegiatan_penilaian.exists():
                return JsonResponse({'status' : 'failed', 'message': 'Data role petugas telah digunakan pada master data penilaian'}, status=200)
            
            check_nilai_mitra = MasterNilaiPetugas.objects.filter(petugas__role = request.POST.get('id'))
            if check_nilai_mitra.exists():
                return JsonResponse({'status' : 'failed', 'message': 'Data role petugas telah digunakan pada master data penilaian'}, status=200)

            form = forms.RoleForm(request.POST, instance=data)
            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"status" : "success", "instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({"status" : "failed", "error": form.errors, "message": "Terjadi Kesalahan"}, status=400)
        return JsonResponse({"status" : "failed", "message": "Terjadi Kesalahan"}, status=400)

class MasterRoleExportClassView(LoginRequiredMixin, View):
    def get(self, request):
    
        resource = MasterRoleResource()
        dataset = resource.export()

        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Jabatan Petugas.xls"'
        return response 