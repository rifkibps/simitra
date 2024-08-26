import json
from pprint import pprint
from django.core.serializers.json import DjangoJSONEncoder

from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.shortcuts import render
from django.views import View
from django.core import serializers
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from . import models
from . import forms
import string
import statistics

from django.urls import reverse_lazy
from django.shortcuts import redirect

from django.http import HttpResponse
from .resources import MasterNilaiResource, MasterKegiatanResource, IndikatorKegiatanResources

from master_petugas.models import RoleMitra, AlokasiPetugas
from master_survey.models import SurveyModel

import itertools
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill
from master_petugas import utils

class PenilaianPetugasClassView(LoginRequiredMixin, View):
    
    def get(self, request):
        context = {
            'title' : 'Kegiatan Penilaian',
            'data_survei' : SurveyModel.objects.all(),
            'data' : models.KegiatanPenilaianModel.objects.all(),
            'roles' : RoleMitra.objects.all(),
            'form' : forms.KegiatanPenilaianForm()
            }
    
        return render(request, 'master_penilaian/kegiatan_penilaian.html', context)


    def post(self, request):

        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.KegiatanPenilaianForm(request.POST)
            if form.is_valid():
                instance = form.save()
                user_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": user_instance, 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)
    

class MasterPenilaianDeleteView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')

                data_petugas = models.KegiatanPenilaianModel.objects.filter(pk = id)
                if data_petugas.exists():

                    check_db = models.MasterNilaiPetugas.objects.filter(penilaian__kegiatan_penilaian = id)
                    if check_db.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data kegiatan penilaian telah digunakan pada master data penilaian'}, status=200)
                
                    check_db_1 = models.IndikatorKegiatanPenilaian.objects.filter(kegiatan_penilaian = id)
                    if check_db_1.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data kegiatan penilaian telah digunakan pada master data penilaian'}, status=200)
                    
                    data_petugas.delete()
                    return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)


class MasterPenilaianDetailView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')
                kegiatan_penilaian = get_object_or_404(models.KegiatanPenilaianModel, pk=id)

                context_data = {}
                context_data['id']              = kegiatan_penilaian.id
                context_data['nama_kegiatan']   = kegiatan_penilaian.nama_kegiatan
                context_data['survey_id']       = str(kegiatan_penilaian.survey.id)
                context_data['tgl_penilaian']   = kegiatan_penilaian.tgl_penilaian
                context_data['status']          = kegiatan_penilaian.status
                context_data['role_permitted']  = list(kegiatan_penilaian.role_permitted.values_list('id', flat=True))

                return JsonResponse({'status' : 'success', 'instance': context_data}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 
    

class MasterPenilaianUpdateView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:

            data = get_object_or_404(models.KegiatanPenilaianModel, pk=request.POST.get('id'))

            form = forms.KegiatanPenilaianForm(request.POST, instance=data)

            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)
    

class PenilaianGetBySurveiClassView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                data = models.KegiatanPenilaianModel.objects.filter(survey=request.POST.get('survey_id')).values()
                return JsonResponse({"instance": list(data)}, status=200)
        
        return JsonResponse({"error": ""}, status=400)
    

class AlokasiGetBySurveiClassView(LoginRequiredMixin, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                penilaian_id = int(request.POST.get('penilaian_id'))
                data = AlokasiPetugas.objects.filter(survey=request.POST.get('survey_id')).values('id', 'petugas__kode_petugas', 'petugas__nama_petugas', 'role', 'role__jabatan' )

                kegiatan_penilaian = get_object_or_404(models.KegiatanPenilaianModel, pk=penilaian_id)
                role_permitted = kegiatan_penilaian.role_permitted.values_list('id', flat=True)

                

                data_2 = models.IndikatorKegiatanPenilaian.objects.filter(kegiatan_penilaian = penilaian_id).values('id', 'indikator_penilaian__nama_indikator' )

                final_data = []
                for dt in data :

                    if dt['role'] in role_permitted:
                        db_check = models.MasterNilaiPetugas.objects.filter(petugas = dt['id'], penilaian__kegiatan_penilaian = penilaian_id)
                        if db_check.exists() == False:
                            final_data.append(dt)

                return JsonResponse({"instance": list(final_data), "instance_2": list(data_2)}, status=200)
        
        return JsonResponse({"error": ""}, status=400)


class MasterPenilaianJsonResponseClassView(LoginRequiredMixin, View):

    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.KegiatanPenilaianModel.objects

        if datatables.get('survei_filter'):
            data = data.filter(survey = datatables.get('survei_filter'))
        
        if datatables.get('status_filter'):
            data = data.filter(status = datatables.get('status_filter'))

        if datatables.get('role_filter'):
            data = data.filter(role_permitted = datatables.get('role_filter'))

        data = data.all().exclude(Q(nama_kegiatan=None)|Q(survey=None)|Q(tgl_penilaian=None)|Q(status=None)|Q(role_permitted=None))
        records_total = data.count()
        records_filtered = records_total
        
        if search:
            data = models.KegiatanPenilaianModel.objects

            if datatables.get('survei_filter'):
                data = data.filter(survey = datatables.get('survei_filter'))
            
            if datatables.get('status_filter'):
                data = data.filter(status = datatables.get('status_filter'))

            if datatables.get('role_filter'):
                data = data.filter(role_permitted = datatables.get('role_filter'))

            data = data.filter(
                Q(nama_kegiatan__icontains=search)|
                Q(survey__nama__icontains=search)|
                Q(tgl_penilaian__icontains=search)|
                Q(status__icontains=search)|
                Q(role_permitted__jabatan__icontains=search)
            ).exclude(Q(nama_kegiatan=None)|Q(survey=None)|Q(sumber=None)|Q(tgl_penilaian=None)|Q(status=None)|Q(role_permitted=None))

            records_total = data.count()
            records_filtered = records_total
        
        data = data.order_by(order_col_name)
        # Conf Paginator
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list
            {"data": 'nama_kegiatan'},
            {"data": 'survey__nama'},
            {"data" : 'tgl_penilaian'},
            {"data" : 'status'},
            {"data" : 'aksi'},
        
        data = [
           
            {
                'nama_kegiatan': obj.nama_kegiatan,
                'survey__nama': obj.survey.nama,
                'tgl_penilaian': obj.tgl_penilaian.strftime('%d-%m-%Y'),
                'role_permitted__jabatan' : ', '.join(obj.role_permitted.values_list('jabatan', flat=True)),
                'status': f'<span class="badge badge-primary-lighten"> {obj.get_status_display()} </span>'  if obj.status == '0' else f'<span class="badge badge-danger-lighten"> {obj.get_status_display()} </span>',
                'aksi': f'<a href="javascript:void(0);" onclick="editKegiatanPenilaian({obj.id})" class="action-icon"><i class="mdi mdi-square-edit-outline"></i></a> <a href="javascript:void(0);" onclick="hapusKegiatanPenilaian({obj.id});" class="action-icon"> <i class="mdi mdi-delete"></i></a>'
            } for obj in object_list
        ]
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }


class MasterPenilaianExportView(LoginRequiredMixin, View):

    def get(self, request):
        resource = MasterKegiatanResource()
        dataset = resource.export()

        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Kegiatan Penilaian.xls"'
        return response 


#INDIKATOR PENILAIAN
class IndikatorPenilaianClassView(LoginRequiredMixin, View):
    
    def get(self, request):
        context = {
            'title' : 'Indikator Penilaian',
            'form' : forms.IndikatorPenilaianForm()
            }

        return render(request, 'master_penilaian/indikator_penilaian.html', context)


    def post(self, request):

        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.IndikatorPenilaianForm(request.POST)
            if form.is_valid():
                instance = form.save()
                user_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": user_instance, 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)
    

class IndiakatorPenilaianJsonResponseClassView(LoginRequiredMixin, View):

    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.IndikatorPenilaian.objects.all().exclude(Q(nama_indikator=None)|Q(deskripsi_penilaian=None))
        records_total = data.count()
        records_filtered = records_total
        
        if search:
            data = models.IndikatorPenilaian.objects.filter(
                Q(nama_indikator__icontains=search)|
                Q(deskripsi_penilaian__icontains=search)
            ).exclude(Q(nama_indikator=None)|Q(deskripsi_penilaian=None))

            records_total = data.count()
            records_filtered = records_total
        
        data = data.order_by(order_col_name)
        # Conf Paginator
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        data = [
            {
                'nama_indikator': obj.nama_indikator,
                'deskripsi_penilaian': obj.deskripsi_penilaian,
                'aksi': f'<button class="btn btn-primary" onclick="editIndikator({obj.id})">Edit</button> <button class="btn btn-danger" onclick="deleteIndikator({obj.id});">Delete</button>'
            } for obj in object_list
        ]
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }


class IndikatorPenilaianDeleteView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')
                
                data_petugas = models.IndikatorPenilaian.objects.filter(pk = id)
                if data_petugas.exists():

                    check_kegiatan_penilaian = models.IndikatorKegiatanPenilaian.objects.filter(indikator_penilaian = id)
                    if check_kegiatan_penilaian.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data Indikator Penilaian telah digunakan pada master data penilaian.'}, status=200)

                    check_db = models.MasterNilaiPetugas.objects.filter(penilaian__indikator_penilaian = id)
                    if check_db.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data Indikator Penilaian telah digunakan pada master data penilaian.'}, status=200)
                
                    data_petugas.delete()
                    return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)
    

class IndikatorPenilaianDetailView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                id = request.POST.get('id')
                
                # check_db = models.MasterNilaiPetugas.objects.filter(penilaian__indikator_penilaian = id)
                # if check_db.exists():
                    # return JsonResponse({'status' : 'failed', 'message': 'Data Indikator Penilaian telah digunakan pada master data penilaian.'}, status=200)
                
                indikator_penilaian = models.IndikatorPenilaian.objects.filter(pk=id)
                if indikator_penilaian.exists():
                    return JsonResponse({'status' : 'success', 'instance': list(indikator_penilaian.values())[0]}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 
    

class IndikatorPenilaianUpdateView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:

            data = get_object_or_404(models.IndikatorPenilaian, pk=request.POST.get('id'))

            # check_db = models.MasterNilaiPetugas.objects.filter(penilaian__indikator_penilaian = request.POST.get('id'))
            # if check_db.exists():
                # return JsonResponse({'status' : 'failed', 'message': 'Data Indikator Penilaian telah digunakan pada master data penilaian.'}, status=200)

            form = forms.IndikatorPenilaianForm(request.POST, instance=data)
            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)


# INDIKATOR KEGIATAN PENILAIAN
class IndikatorKegiatanPenilaianClassView(LoginRequiredMixin, View):

    def get(self, request):
        context = {
            'title' : 'Kegiatan Penilaian',
            'data_kegiatan': models.KegiatanPenilaianModel.objects.all(),
            'data_indikator': models.IndikatorPenilaian.objects.all(),
            'form' : forms.IndikatorKegiatanPenilaianForm()
            }
        
        return render(request, 'master_penilaian/indikator_kegiatan_penilaian.html', context)


    def post(self, request):

        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.IndikatorKegiatanPenilaianForm(request.POST)
           
            if form.is_valid():

                check_db2 = models.IndikatorKegiatanPenilaian.objects.filter(kegiatan_penilaian = form.data['kegiatan_penilaian'], indikator_penilaian = form.data['indikator_penilaian'])
                
                print('asd', check_db2.exists())
                if check_db2.exists():

                    return JsonResponse({'status' : 'failed', 'message': 'Data survei telah tersedia pada database.'}, status=200)
                else:
                    instance = form.save()
                    user_instance = serializers.serialize('json', [ instance, ])
                    
                    # send to client side.
                    return JsonResponse({'status' : 'success', "instance": user_instance, 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                print('GAK LOLOS FORM')
                return JsonResponse({'status' : 'failed', "error": form.errors}, status=400)
            
        return JsonResponse({'status' : 'failed', "error": ""}, status=400)
    


class IndikatorKegiatanPenilaianJsonResponseClassView(LoginRequiredMixin, View):

    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.IndikatorKegiatanPenilaian.objects

        if datatables.get('kegiatan_filter'):
            data = data.filter(kegiatan_penilaian = datatables.get('kegiatan_filter'))
        
        if datatables.get('indikator_filter'):
            data = data.filter(indikator_penilaian = datatables.get('indikator_filter'))

        data = data.all().exclude(Q(kegiatan_penilaian=None)|Q(indikator_penilaian=None))
        records_total = data.count()
        records_filtered = records_total
        
        if search:

            data = models.IndikatorKegiatanPenilaian.objects

            if datatables.get('kegiatan_filter'):
                data = data.filter(kegiatan_penilaian = datatables.get('kegiatan_filter'))
            
            if datatables.get('indikator_filter'):
                data = data.filter(indikator_penilaian = datatables.get('indikator_filter'))

            data = data.filter(
                Q(kegiatan_penilaian__survey__nama=search)|
                Q(kegiatan_penilaian__nama_kegiatan__icontains=search)|
                Q(indikator_penilaian__nama_indikator__icontains=search)
            ).exclude(Q(kegiatan_penilaian=None)|Q(indikator_penilaian=None))

            records_total = data.count()
            records_filtered = records_total
        
        data = data.order_by(order_col_name)
        # Conf Paginator
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list
         
        data = [
            {
                'kegiatan_penilaian__nama_kegiatan': obj.kegiatan_penilaian.nama_kegiatan,
                'indikator_penilaian__nama_indikator': obj.indikator_penilaian.nama_indikator,
                'aksi': f'<a href="javascript:void(0);" onclick="editIndikatorKegiatan({obj.id})" class="action-icon"><i class="mdi mdi-square-edit-outline"></i></a> <a href="javascript:void(0);" onclick="deleteIndikatorKegiatan({obj.id});" class="action-icon"> <i class="mdi mdi-delete"></i></a>'
            } for obj in object_list
        ]
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }


class IndikatorKegiatanPenilaianDeleteView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')
               
                data_petugas = models.IndikatorKegiatanPenilaian.objects.filter(pk = id)
                if data_petugas.exists():

                    check_db = models.MasterNilaiPetugas.objects.filter(penilaian = id)
                    if check_db.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data survei telah digunakan pada master data penilaian.'}, status=200)

                    data_petugas.delete()
                    return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)
    

class IndikatorKegiatanPenilaianDetailView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                id = request.POST.get('id')

                indikator_penilaian = models.IndikatorKegiatanPenilaian.objects.filter(pk=id)

                if indikator_penilaian.exists():

                    check_db = models.MasterNilaiPetugas.objects.filter(penilaian = id)
                    if check_db.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data survei telah digunakan pada master data penilaian.'}, status=200)
                
                    return JsonResponse({'status' : 'success', 'instance': list(indikator_penilaian.values())[0]}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 
    

class IndikatorKegiatanPenilaianUpdateView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:

            data = get_object_or_404(models.IndikatorKegiatanPenilaian, pk=request.POST.get('id'))

            form = forms.IndikatorKegiatanPenilaianForm(request.POST, instance=data)
            if form.is_valid():
                
                check_db = models.MasterNilaiPetugas.objects.filter(penilaian = form.data['id'])
                if check_db.exists():
                    return JsonResponse({'status' : 'failed', 'message': 'Data survei telah digunakan pada master data penilaian.'}, status=200)
                
                check_db2 = models.IndikatorKegiatanPenilaian.objects.filter(~Q(id = form.data['id'] ) & Q(kegiatan_penilaian = form.data['kegiatan_penilaian'] ) & Q(indikator_penilaian = form.data['indikator_penilaian']) )
                if check_db2.exists():
                    return JsonResponse({'status' : 'failed', 'message': 'Data survei telah tersedia pada database.'}, status=200)
                
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({'status' : 'success', "instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({'status' : 'failed', "error": form.errors}, status=400)
        return JsonResponse({'status' : 'failed', "error": ""}, status=400)


class IndikatorKegiatanPenilaianExportView(LoginRequiredMixin, View):

    def get(self, request):
    
        resource = IndikatorKegiatanResources()
        dataset = resource.export()

        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Indikator Kegiatan Penilaian.xls"'
        return response
    

# NILAI MITRA
class NilaiMitraClassView(LoginRequiredMixin, View):
    
    def get(self, request):
        context = {
            'title' : 'Nilai Mitra',
            'data_mitra' : AlokasiPetugas.objects.all(),
            'data_role' : RoleMitra.objects.all(),
            'data_survei' : SurveyModel.objects.all(),
            'data_indikator_penilaian' : models.IndikatorPenilaian.objects.all(),
            'data_kegiatan_penilaian' : models.KegiatanPenilaianModel.objects.all(),
            'form' : forms.PenilaianMitraForm(),
            'form_upload' : forms.NilaiFormUpload()
            }

        return render(request, 'master_penilaian/nilai-mitra.html', context)


    def post(self, request):

        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:

            form = forms.PenilaianMitraForm(request.POST)

            if form.is_valid():            
                model = models.MasterNilaiPetugas

                df = form.cleaned_data

                objs_create = []
                objs_update = []

                for dt in df:
                    
                    petugas = AlokasiPetugas.objects.get(pk = dt['petugas'])
                    penilaian = models.IndikatorKegiatanPenilaian.objects.get(pk = dt['penilaian'])

                    role_permitted = list(penilaian.kegiatan_penilaian.role_permitted.values_list('id', flat=True))
                    if petugas.role.id not in role_permitted:
                        return JsonResponse({"status": "failed", "messages":  f'Data alokasi petugas dengan role {petugas.role.jabatan} tidak diizinkan untuk mengikuti kegiatan penilaian "<i>{penilaian.kegiatan_penilaian.nama_kegiatan}</i>"'})
                    
                    nilai = dt['nilai']
                    catatan = dt['catatan']

                    db_check = model.objects.filter(petugas = petugas, penilaian = penilaian)

                    if db_check.exists():
                        nilai_mitra_update = db_check.first()
                        nilai_mitra_update.nilai = nilai
                        nilai_mitra_update.catatan = catatan
                        objs_update.append(nilai_mitra_update)

                    else:
                        objs_create.append(
                            model(
                                petugas = petugas,
                                penilaian = penilaian,
                                nilai = nilai,
                                catatan = catatan,
                            )
                        )

                msg = ''
                if len(objs_create) > 0:
                    model.objects.bulk_create(objs_create)
                    msg += f"Data <strong>berhasil</strong> ditambahkan.<br>"

                if len(objs_update) > 0:
                    model.objects.bulk_update(objs_update, ['nilai', 'catatan'])
                    msg += f"Data <strong>berhasil</strong> diperbarui.<br>"

                return JsonResponse({"status": "success", "messages":  msg})
            
            return JsonResponse({"error": form.errors}, status=400)
        else:
            return JsonResponse({"error": ""}, status=400)


class NilaiMitraJsonResponseClassView(LoginRequiredMixin, View):

    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.MasterNilaiPetugas.objects

        if datatables.get('mitra_filter'):
            data = data.filter(petugas__petugas = datatables.get('mitra_filter'))
        
        if datatables.get('role_filter'):
            data = data.filter(petugas__role = datatables.get('role_filter'))

        if datatables.get('survei_filter'):
            data = data.filter(penilaian__kegiatan_penilaian__survey = datatables.get('survei_filter'))

        if datatables.get('kegiatan_filter'):
            data = data.filter(penilaian__kegiatan_penilaian = datatables.get('kegiatan_filter'))

        data = data.all().exclude(Q(petugas=None)|Q(penilaian=None)|Q(nilai=None)|Q(catatan=None))
        records_total = data.count()
        records_filtered = records_total
    
        if search:

            data = models.MasterNilaiPetugas.objects

            if datatables.get('mitra_filter'):
                data = data.filter(petugas__petugas = datatables.get('mitra_filter'))

            if datatables.get('role_filter'):
                data = data.filter(petugas__role = datatables.get('role_filter'))

            if datatables.get('survei_filter'):
                data = data.filter(penilaian__kegiatan_penilaian__survey = datatables.get('survei_filter'))

            if datatables.get('kegiatan_filter'):
                data = data.filter(penilaian__kegiatan_penilaian = datatables.get('kegiatan_filter'))

            data = data.filter(
                Q(petugas__petugas__kode_petugas__icontains=search)|
                Q(petugas__petugas__nama_petugas__icontains=search)|
                Q(petugas__role__jabatan__icontains=search)|
                Q(penilaian__kegiatan_penilaian__nama_kegiatan__icontains=search)|
                Q(penilaian__indikator_penilaian__nama_indikator=search)|
                Q(nilai__icontains=search)|
                Q(catatan__icontains=search)
            ).exclude(Q(petugas=None)|Q(penilaian=None)|Q(nilai=None)|Q(catatan=None))

            records_total = data.count()
            records_filtered = records_total
        
        data = data.order_by(order_col_name)
        # Conf Paginator
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        data = [
            {
                'petugas__petugas__kode_petugas': obj.petugas.petugas.kode_petugas,
                'petugas__petugas__nama_petugas': f'<a href="{reverse_lazy("master_petugas:mitra-view-detail", kwargs={"mitra_id": obj.petugas.petugas.id})}" class="text-body" target="_blank">{obj.petugas.petugas.nama_petugas}</a>',
                'petugas__role__jabatan': obj.petugas.role.jabatan,
                'penilaian__kegiatan_penilaian__nama_kegiatan': obj.penilaian.kegiatan_penilaian.nama_kegiatan,
                'penilaian__indikator_penilaian__nama_indikator': obj.penilaian.indikator_penilaian.nama_indikator,
                'nilai': obj.nilai,
                'catatan' :  obj.catatan,
            } for obj in object_list
        ]
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }


class NilaiMitraDetailClassView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                id = request.POST.get('id')

                nilai_mitra = models.MasterNilaiPetugas.objects.filter(pk=id)

                if nilai_mitra.exists():
                    return JsonResponse({'status' : 'success', 'instance': list(nilai_mitra.values())[0]}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 


class NilaiMitraUpdateClassView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:

            data = get_object_or_404(models.MasterNilaiPetugas, pk=request.POST.get('id'))
            
            form = forms.PenilaianMitraForm(request.POST, instance=data)
            if form.is_valid():

                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)


class NilaiMitraDeleteClassView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id_kegiatan = request.POST.get('id_kegiatan')
                id_alokasi = request.POST.get('id_alokasi')
                nilai_mitra = models.MasterNilaiPetugas.objects.filter(petugas = id_alokasi, penilaian__kegiatan_penilaian=id_kegiatan)

                if nilai_mitra.exists() == False:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                 
                nilai_mitra.delete()
                return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)
    

class NilaiMitraExportClassView(LoginRequiredMixin, View):

    def get(self, request):
    
        resource = MasterNilaiResource()
        dataset = resource.export()

        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Master Nilai Mitra.xls"'
        return response 
    

class NilaiMitraTemplateClassView(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs): 

        kegiatan_penilaian = self.kwargs['kegiatan']

        kegiatan = models.KegiatanPenilaianModel.objects.filter(pk=kegiatan_penilaian)

        if kegiatan.exists() == False:
            return redirect(reverse_lazy('master_penilaian:nilai-mitra'))
        
        kegiatan = kegiatan.first()

        survey_id = kegiatan.survey.id
        alokasi_petugas = AlokasiPetugas.objects.filter(survey = survey_id)

        indikator_kegiatan_penilaian = models.IndikatorKegiatanPenilaian.objects.filter(kegiatan_penilaian = kegiatan.id)

        wb = Workbook()

        ws = wb.active

        # Ini untuk header columns
        ws.title = 'Upload Nilai Mitra Petugas'

        indikator = list(indikator_kegiatan_penilaian.values_list('indikator_penilaian__nama_indikator', flat=True))
        catatan_indikator = [f"Catatan Personal {dt}" for dt in indikator]
        indikator_penilaian = [f"Penilaian Indikator {dt}" for dt in indikator]
        headers = ['No', 'ID Alokasi Mitra', 'ID Kegiatan Penilaian', 'Kode Petugas', 'Nama Petugas', 'Survei', 'Jabatan', 'Kegiatan Penilaian'] + indikator_penilaian + catatan_indikator


        head_row = 2
        alphabet_string = utils.generate_headers_excel(100)
        
        header_cols = np.array(ws[f'A{head_row}':f'{alphabet_string[len(headers)]}{head_row}'])
        # Set value and style for header
        for v,c in zip(headers, header_cols.T.flatten()):
           
            # Set style
            c.font = Font(name='Calibri', size=12)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type = "solid")
            c.value = v


        # Adjustment cols
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
            
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 22.50
    
        ws.merge_cells(f'A1:{alphabet_string[8:8 + len(indikator_penilaian)*2][-1]}1')
        ws['A1'] = 'Template Upload Nilai Mitra'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Calibri',bold=True, size=14)

        no_ = 0
        for idx, dt_alokasi in enumerate(alokasi_petugas):

            db_query_check = models.MasterNilaiPetugas.objects.filter(petugas = dt_alokasi.id)

            if db_query_check.exists():
                db_query_check2 = db_query_check.filter(Q(penilaian__kegiatan_penilaian = kegiatan.id))

                if db_query_check2.exists():
                    ws[f'A{no_+3}'] = no_ + 1
                    ws[f'B{no_+3}'] = dt_alokasi.id
                    ws[f'C{no_+3}'] = kegiatan.id
                    ws[f'D{no_+3}'] = dt_alokasi.petugas.kode_petugas
                    ws[f'E{no_+3}'] = dt_alokasi.petugas.nama_petugas
                    ws[f'F{no_+3}'] = dt_alokasi.survey.nama
                    ws[f'G{no_+3}'] = dt_alokasi.role.jabatan
                    ws[f'H{no_+3}'] = kegiatan.nama_kegiatan

                    for c in alphabet_string[:8]:
                        ws[f'{c}{no_+3}'].fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type = "solid")

                    for idx2, db_row in enumerate(db_query_check):

                        db_row_indikator = db_row.penilaian.indikator_penilaian.nama_indikator
            
                        indikator_index = indikator.index(db_row_indikator) if db_row_indikator in indikator else -1

                        if indikator_index >= 0:
                            ws[f'{alphabet_string[8 + indikator_index]}{no_+3}'] = db_row.nilai
                            ws[f'{alphabet_string[8 + len(indikator) +idx2]}{no_+3}'] = db_row.catatan
                            ws[f'{alphabet_string[8 + indikator_index]}{no_+3}'].fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type = "solid")
                            ws[f'{alphabet_string[8 + len(indikator) +idx2]}{no_+3}'].fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type = "solid")
                    no_ += 1
            else:
                
                role_permitted = list(kegiatan.role_permitted.values_list('id', flat=True))

                if dt_alokasi.role.id in role_permitted:
                    ws[f'A{no_+3}'] = no_ + 1
                    ws[f'B{no_+3}'] = dt_alokasi.id
                    ws[f'C{no_+3}'] = kegiatan.id
                    ws[f'D{no_+3}'] = dt_alokasi.petugas.kode_petugas
                    ws[f'E{no_+3}'] = dt_alokasi.petugas.nama_petugas
                    ws[f'F{no_+3}'] = dt_alokasi.survey.nama
                    ws[f'G{no_+3}'] = dt_alokasi.role.jabatan
                    ws[f'H{no_+3}'] = kegiatan.nama_kegiatan

                    no_ += 1

        ws.column_dimensions['B'].hidden= True
        ws.column_dimensions['C'].hidden= True
        # ws[f'A{no_+4}'] = 'Keterangan:'
        # ws[f'D{no_+5}'] = 'Telah tercatat pada database, ubah data untuk mengupdate nilai mitra.'
        # ws[f'A{no_+5}'].fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type = "solid")
        # ws[f'A{no_+4}'].font = Font(name='Calibri', size=10)
        # ws[f'D{no_+5}'].font = Font(name='Calibri', size=10)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Template Upload Nilai Mitra.xlsx'

        wb.save(response)
        return response


class NilaiMitraUploadClassView(LoginRequiredMixin, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
           
            form = forms.NilaiFormUpload(request.POST, request.FILES)
            model = models.MasterNilaiPetugas

            if form.is_valid():
                
                df = form.cleaned_data

                objs_create = []
                objs_update = []

  
                for dt in df:
                    
                    petugas = AlokasiPetugas.objects.get(pk = dt['petugas'])
                    penilaian = models.IndikatorKegiatanPenilaian.objects.get(pk = dt['penilaian'])

                    role_permitted = list(penilaian.kegiatan_penilaian.role_permitted.values_list('id', flat=True))
                    if petugas.role.id not in role_permitted:
                        continue

                    nilai = dt['nilai']
                    catatan = dt['catatan']

                    db_check = model.objects.filter(petugas = petugas, penilaian = penilaian)

                    if db_check.exists():
                        nilai_mitra_update = db_check.first()
                        nilai_mitra_update.nilai = nilai
                        nilai_mitra_update.catatan = catatan

                        objs_update.append(nilai_mitra_update)
                    else:
                        objs_create.append(
                            model(
                                petugas = petugas,
                                penilaian = penilaian,
                                nilai = nilai,
                                catatan = catatan,
                            )
                        )

                msg = ''
                if len(objs_create) > 0:
                    model.objects.bulk_create(objs_create)
                    msg += f"<strong>{len(objs_create)}</strong> Data berhasil diupload.<br>"

                if len(objs_update) > 0:
                    model.objects.bulk_update(objs_update, ['nilai', 'catatan'])
                    msg += f"<strong>{len(objs_update)}</strong> Data berhasil diperbarui."

                return JsonResponse({"status": "success", "messages":  msg})
            else:
                
                error_messages = list(itertools.chain.from_iterable(form.errors['import_file'].as_data()))
                return JsonResponse({"status": "error", "messages": error_messages})

        return JsonResponse({"error": ""}, status=403)  


class GenerateTableNilaiClassView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                kegiatan_penilaian = request.POST.get('kegiatan_penilaian')

                kegiatan = models.KegiatanPenilaianModel.objects.filter(pk=kegiatan_penilaian)
                if kegiatan.exists() == False:
                    return JsonResponse({'status': 'failed', 'data': 'Data tidak tersedia'})
                
                kegiatan = kegiatan.first()
                survey_id = kegiatan.survey.id

                if request.POST.get('filter_role_nilai_mitra'):
                    alokasi_petugas = AlokasiPetugas.objects.filter(survey = survey_id, role = request.POST.get('filter_role_nilai_mitra'))
                else:
                    alokasi_petugas = AlokasiPetugas.objects.filter(survey = survey_id)

                master_nilai = models.MasterNilaiPetugas.objects.filter(penilaian__kegiatan_penilaian = kegiatan_penilaian)
                indikator_penilaian = list(master_nilai.values_list('penilaian__indikator_penilaian__nama_indikator', flat=True).distinct())
                indikator_col = [ f"Penilaian {indikator}" for indikator in indikator_penilaian]
                catatan_col = [ f"Catatan {indikator}" for indikator in indikator_penilaian]
                
                headers = ['ID Mitra', 'Mitra',	'Jabatan', 'Kegiatan Penilaian']
                headers += indikator_col + ['Rerata'] + catatan_col + ['Aksi']
                
                thead = '<tr>'
                for header in headers:
                    thead += f'<th>{header}</th>'
                thead += '</tr>'

                tbody_data = []
                for idx, dt_alokasi in enumerate(alokasi_petugas):

                    dt_row = []
                    db_query_check = models.MasterNilaiPetugas.objects.filter(petugas = dt_alokasi.id, penilaian__kegiatan_penilaian = kegiatan.id)
                    dt_row.append(dt_alokasi.id) #[0]
                    dt_row.append(dt_alokasi.petugas.id) #[0]
                    dt_row.append(dt_alokasi.petugas.kode_petugas) #[1]
                    dt_row.append(dt_alokasi.petugas.nama_petugas) #[2]
                    dt_row.append(dt_alokasi.role.jabatan) #[3]
                    dt_row.append(kegiatan.nama_kegiatan) #[4]

                    if db_query_check.exists():
                        rerata = []
                        for idx2, db_row in enumerate(db_query_check):

                            db_row_indikator = db_row.penilaian.indikator_penilaian.nama_indikator
                
                            indikator_index = indikator_penilaian.index(db_row_indikator) if db_row_indikator in indikator_penilaian else -1
                            
                            if indikator_index >= 0:
                                rerata.append(db_row.nilai)
                                dt_row.insert(5+indikator_index+1, db_row.nilai)
                                dt_row.insert(5+len(indikator_penilaian)+indikator_index+2, db_row.catatan)

                        rerata = round(statistics.mean(rerata), 2)
                        dt_row.insert(5+len(indikator_penilaian)+1, rerata)
                        tbody_data.append(dt_row)

                pprint(tbody_data)
                tbody = ''
                for data in tbody_data:
                    tbody += '<tr>'
                    for idx, dt in enumerate(data):
                        if idx == 0:
                            button_action = f'<td><button class="btn btn-primary" onclick="updateNilaiMitra({dt})" >Edit</button> <button class="btn btn-danger" onclick="deleteNilaiMitra({dt})">Hapus</button></td>'
                        elif idx == 1:
                            continue
                        elif idx == 3:
                            tbody += f'<td><a href="{reverse_lazy("master_petugas:mitra-view-detail", kwargs={"mitra_id": data[1]})}" class="text-body" target="_blank">{data[3]}</a></td>'
                        else:
                            tbody += f'<td>{dt}</td>'

                    tbody += button_action
                    tbody += '</tr>'

                
                return JsonResponse({'status' : 'success', 'data': {'thead' : thead, 'tbody' : tbody}}, status=200)

        return JsonResponse({'status': 'Invalid request'}, status=400)


class GetNilaiMitraClassView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                id_alokasi = request.POST.get('id_alokasi')
                id_kegiatan = request.POST.get('id_kegiatan')
                data_mitra = AlokasiPetugas.objects.values('id', 'petugas__kode_petugas', 'petugas__nama_petugas', 'role__jabatan', 'survey__id', 'survey__nama').get(pk=id_alokasi)
                data_kegiatan = models.KegiatanPenilaianModel.objects.values('id', 'nama_kegiatan').get(pk = id_kegiatan)
                nilai_mitra = models.MasterNilaiPetugas.objects.filter(petugas = id_alokasi, penilaian__kegiatan_penilaian=id_kegiatan).values('penilaian', 'penilaian__indikator_penilaian__nama_indikator', 'nilai' , 'catatan' )

                if nilai_mitra.exists():
                    return JsonResponse({'status': 'success', 'data_kegiatan': data_kegiatan, 'data_mitra': data_mitra, 'nilai_mitra': list(nilai_mitra)}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 
